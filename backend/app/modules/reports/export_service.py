from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.transaction import TransactionType
from app.modules.reports.service import ReportCategoryTotal, ReportData, ReportTransaction

CurrencyCode = str

MONTH_NAMES = [
    "Ocak",
    "Şubat",
    "Mart",
    "Nisan",
    "Mayıs",
    "Haziran",
    "Temmuz",
    "Ağustos",
    "Eylül",
    "Ekim",
    "Kasım",
    "Aralık",
]

CURRENCY_SYMBOLS = {
    "TRY": "₺",
    "USD": "$",
    "EUR": "€",
}

CURRENCY_EXCEL_FORMATS = {
    "TRY": '"₺"#,##0.00',
    "USD": '"$"#,##0.00',
    "EUR": '"€"#,##0.00',
}

FONT_NAME = "Helvetica"


def normalize_currency(currency: str | None) -> str:
    value = (currency or "TRY").upper()
    return value if value in CURRENCY_SYMBOLS else "TRY"


def register_pdf_font() -> str:
    global FONT_NAME
    if FONT_NAME != "Helvetica":
        return FONT_NAME

    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/local/share/fonts/dejavu/DejaVuSans.ttf"),
    ]
    for font_path in candidates:
        if font_path.exists():
            pdfmetrics.registerFont(TTFont("CredentiaSans", str(font_path)))
            FONT_NAME = "CredentiaSans"
            break
    return FONT_NAME


def period_label(report: ReportData) -> str:
    return f"{MONTH_NAMES[report.month - 1]} {report.year}"


def file_period(report: ReportData) -> str:
    return f"{report.year}-{report.month:02d}"


def format_currency(value: float, currency: str) -> str:
    symbol = CURRENCY_SYMBOLS[normalize_currency(currency)]
    safe_value = value if isinstance(value, (int, float)) else 0
    formatted = f"{safe_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{symbol}{formatted}"


def format_date(value) -> str:
    return value.strftime("%d.%m.%Y")


def transaction_type_label(value: TransactionType) -> str:
    return "Gelir" if value == TransactionType.income else "Gider"


def table_or_empty(rows: list[list[object]], col_widths: list[float]) -> Table | Paragraph:
    if len(rows) <= 1:
        return Paragraph("Bu dönemde kayıtlı veri bulunamadı.", get_paragraph_styles()["Body"])

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), register_pdf_font()),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    return table


def get_paragraph_styles() -> dict[str, ParagraphStyle]:
    font_name = register_pdf_font()
    base = getSampleStyleSheet()
    return {
        "Title": ParagraphStyle(
            "CredentiaTitle",
            parent=base["Title"],
            fontName=font_name,
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=8,
        ),
        "Heading": ParagraphStyle(
            "CredentiaHeading",
            parent=base["Heading2"],
            fontName=font_name,
            fontSize=12,
            leading=15,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=10,
            spaceAfter=6,
        ),
        "Body": ParagraphStyle(
            "CredentiaBody",
            parent=base["BodyText"],
            fontName=font_name,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
        ),
        "Small": ParagraphStyle(
            "CredentiaSmall",
            parent=base["BodyText"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#64748b"),
        ),
    }


def category_rows(items: list[ReportCategoryTotal], currency: str) -> list[list[object]]:
    styles = get_paragraph_styles()
    rows: list[list[object]] = [["Kategori", "Tutar"]]
    rows.extend([[Paragraph(item.category, styles["Body"]), format_currency(item.amount, currency)] for item in items])
    return rows


def transaction_rows(items: list[ReportTransaction], currency: str) -> list[list[object]]:
    styles = get_paragraph_styles()
    rows: list[list[object]] = [["Tarih", "Açıklama", "Kategori", "Tür", "Tutar"]]
    rows.extend(
        [
            [
                format_date(item.occurred_on),
                Paragraph(item.description, styles["Body"]),
                Paragraph(item.category, styles["Body"]),
                transaction_type_label(item.type),
                format_currency(item.amount, currency),
            ]
            for item in items
        ]
    )
    return rows


def build_pdf(report: ReportData, currency: str) -> bytes:
    normalized_currency = normalize_currency(currency)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Credentia Finans Raporu {file_period(report)}",
    )
    styles = get_paragraph_styles()
    story: list[object] = [
        Paragraph("Credentia Finans Raporu", styles["Title"]),
        Paragraph(f"Dönem: {period_label(report)}", styles["Body"]),
        Spacer(1, 6),
        Paragraph("Bölüm 1 - Finansal Özet", styles["Heading"]),
        table_or_empty(
            [
                ["Alan", "Değer"],
                ["Toplam gelir", format_currency(report.summary.income, normalized_currency)],
                ["Toplam gider", format_currency(report.summary.expense, normalized_currency)],
                ["Net bakiye", format_currency(report.summary.balance, normalized_currency)],
            ],
            [55 * mm, 65 * mm],
        ),
        Paragraph("Bölüm 2 - Kategori Bazlı Gelirler", styles["Heading"]),
        table_or_empty(category_rows(report.income_categories, normalized_currency), [95 * mm, 45 * mm]),
        Paragraph("Bölüm 3 - Kategori Bazlı Giderler", styles["Heading"]),
        table_or_empty(category_rows(report.expense_categories, normalized_currency), [95 * mm, 45 * mm]),
        Paragraph("Bölüm 4 - İşlem Dökümü", styles["Heading"]),
        table_or_empty(
            transaction_rows(report.transactions, normalized_currency),
            [24 * mm, 48 * mm, 36 * mm, 22 * mm, 32 * mm],
        ),
        Spacer(1, 8),
        Paragraph("Bu rapor Credentia tarafından kayıtlı finans verilerine göre oluşturulmuştur.", styles["Small"]),
    ]
    document.build(story)
    return buffer.getvalue()


def append_empty_message(sheet) -> None:
    sheet.append(["Bu dönemde kayıtlı veri bulunamadı."])


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 42)


def build_excel(report: ReportData, currency: str) -> bytes:
    normalized_currency = normalize_currency(currency)
    money_format = CURRENCY_EXCEL_FORMATS[normalized_currency]
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Özet"
    summary_sheet.append(["Alan", "Değer"])
    summary_sheet.append(["Dönem", period_label(report)])
    summary_sheet.append(["Toplam gelir", report.summary.income])
    summary_sheet.append(["Toplam gider", report.summary.expense])
    summary_sheet.append(["Net bakiye", report.summary.balance])
    style_header(summary_sheet)
    for row in range(3, 6):
        summary_sheet[f"B{row}"].number_format = money_format
    autosize_columns(summary_sheet)

    transactions_sheet = workbook.create_sheet("İşlemler")
    transactions_sheet.append(["Tarih", "Açıklama", "Kategori", "Tür", "Tutar"])
    if report.transactions:
        for item in report.transactions:
            transactions_sheet.append(
                [
                    item.occurred_on,
                    item.description,
                    item.category,
                    transaction_type_label(item.type),
                    item.amount,
                ]
            )
        for row in range(2, transactions_sheet.max_row + 1):
            transactions_sheet[f"A{row}"].number_format = "dd.mm.yyyy"
            transactions_sheet[f"E{row}"].number_format = money_format
    else:
        append_empty_message(transactions_sheet)
    style_header(transactions_sheet)
    autosize_columns(transactions_sheet)

    categories_sheet = workbook.create_sheet("Kategoriler")
    categories_sheet.append(["Kategori", "Tür", "Toplam tutar"])
    categories = [*report.income_categories, *report.expense_categories]
    if categories:
        for item in categories:
            categories_sheet.append([item.category, transaction_type_label(item.type), item.amount])
        for row in range(2, categories_sheet.max_row + 1):
            categories_sheet[f"C{row}"].number_format = money_format
    else:
        append_empty_message(categories_sheet)
    style_header(categories_sheet)
    autosize_columns(categories_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
